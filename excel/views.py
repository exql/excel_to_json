from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse

from django.contrib.auth import authenticate, login
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from datetime import datetime
from tempfile import NamedTemporaryFile
from .service import * 
from .forms import *
import os
import io
import json
from io import BytesIO


from django.contrib import messages
from django.contrib.auth.models import User, Group
from .models import PerfilUsuario, DatosLab, LabEnsayo
import traceback
from django.db import connection
from excel.utils import convertir_tipo 
from openpyxl.styles import PatternFill






# Create your views here.

# Views de las diferntes paginas
@login_required
def index(request):
    return render(request, 'index.html')
    #return HttpResponse('funciona por ahora')
@login_required
def planes(request):
    return render(request, 'planes.html')

def aie(request):
    return render(request, 'aie.html')

def brucelosis(request):
    return render(request, 'brucelosis.html')

def triqui(request):
    return render(request, 'triqui.html')

def contacto(request):
    return render(request, 'contacto.html')

def nosotros(request):
    return render(request, 'nosotros.html')

def preguntas(request):
    return render(request, 'preguntas.html')

def instructivos(request):
    return render(request, 'instructivos.html')

def tutoriales(request):
    return render(request, 'tutoriales.html')

def converter(request):
    return render(request, 'converter.html')


def actadigital_AIE(request): # Hoja PDF a Excel para AIE
    return render(request, 'actadigital_AIE.html')

def actadigital_bru(request):
    return render(request, 'actadigital_bru.html')
 # reemplazar este por el actadigital_bru.html

def actadigital_bru2(request):
    return render(request, 'actadigital_bru_2.html')
 # reemplazar este por el actadigital_bru.html

def actadigital_bru3(request):
    return render(request, 'actadigital_bru3.html')
 # reemplazar este por el actadigital_bru.html

def actadigital_AIE2(request): # Hoja PDF a Excel para AIE
    return render(request, 'actadigital_AIE2.html')

def Actadigital_Aujeszky(request):
    # Definimos la lista para el bucle
    contexto = {'actas': [1, 2, 3, 4]}
    return render(request, 'Actadigital_Aujeszky.html', contexto)


#Registrar Usuarios
@login_required
def registrar_usuario(request):
    if request.method == "POST":
        form = RegistroUsuarioForm(request.POST)
        if form.is_valid():
            usuario=form.save()
            return redirect("asignar_laboratorio",user_id=usuario.id)
    else:
        form = RegistroUsuarioForm()

    return render(request, "registro_usuario.html", {"form": form})  # ✅ Ya no usa `excel/`




@login_required
def asignar_laboratorio(request, user_id):
    usuario = get_object_or_404(User, id=user_id)

    if request.method == "POST":
        lab_form = AsignarLaboratorioForm(request.POST)
        ensayo_formset = AsignarEnsayoFormSet(request.POST)

        if lab_form.is_valid() and ensayo_formset.is_valid():
            laboratorio = lab_form.save(commit=True)

            # ✅ Asociar laboratorio al perfil de usuario
            perfil_usuario, created = PerfilUsuario.objects.get_or_create(usuario=usuario)
            perfil_usuario.datos_lab = laboratorio
            perfil_usuario.save()

            # ✅ Asignar grupo "laboratorio" al usuario
            grupo_lab = Group.objects.get(name="laboratorio")
            usuario.groups.add(grupo_lab)

            # ✅ Crear las asignaciones de ensayos
            for form in ensayo_formset:
                if form.cleaned_data and not form.cleaned_data.get("DELETE", False):
                    ensayo = form.cleaned_data["ensayo"]
                    codigo = form.cleaned_data["codigo_ensayo"]
                    LabEnsayo.objects.create(
                        laboratorio=laboratorio,
                        ensayo=ensayo,
                        codigo_ensayo=codigo
                    )

            return redirect("usuarios_lista")

    else:
        lab_form = AsignarLaboratorioForm()
        ensayo_formset = AsignarEnsayoFormSet()

    return render(request, "asignar_laboratorio.html", {
        "form": lab_form,
        "ensayo_formset": ensayo_formset,
        "usuario": usuario
    })

@csrf_exempt
@login_required
def cambiar_contraseña(request, user_id):
    usuario = get_object_or_404(User, id=user_id)

    try:
        perfil = PerfilUsuario.objects.get(usuario=usuario)  # ✅ Accede correctamente al perfil
    except PerfilUsuario.DoesNotExist:
        perfil = None  # ✅ Evita errores si el perfil no existe

    if request.method == "POST":
        nueva_password = request.POST.get("password")
        if nueva_password:
            usuario.set_password(nueva_password)  # ✅ Guarda la nueva contraseña cifrada
            usuario.save()

            if perfil:
                perfil.password_visible = nueva_password  # ✅ Guarda la clave temporal en el perfil
                perfil.save()

    return redirect("usuarios_lista")  # ✅ Redirige de nuevo a la lista de usuarios



def usuarios_lista(request):
    perfiles = PerfilUsuario.objects.select_related("usuario", "datos_lab").all()

    # Leer datos de la vista SQL
    ensayos_por_lab = {}
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT laboratorio_id, analito, nombreMatriz, tecnica, codigo_rubro
            FROM vista_laboratorio_ensayos
        """)
        for row in cursor.fetchall():
            lab_id = row[0]
            ensayo_info = {
                "analito": row[1],
                "matriz": row[2],
                "tecnica": row[3],
                "codigo_rubro": row[4],
            }
            ensayos_por_lab.setdefault(lab_id, []).append(ensayo_info)

    return render(request, "usuarios_lista.html", {
        "perfiles": perfiles,
        "ensayos_por_lab": ensayos_por_lab
    })




def editar_ensayos_laboratorio(request, lab_id):
    laboratorio = get_object_or_404(DatosLab, id=lab_id)

    if request.method == "POST":
        formset = EditarEnsayosFormSet(request.POST)
        if formset.is_valid():
            LabEnsayo.objects.filter(laboratorio=laboratorio).delete()
            for form in formset:
                if form.cleaned_data and not form.cleaned_data.get("DELETE", False):
                    ensayo = form.cleaned_data["ensayo"]
                    codigo = form.cleaned_data["codigo_ensayo"]
                    LabEnsayo.objects.create(
                        laboratorio=laboratorio,
                        ensayo=ensayo,
                        codigo_ensayo=codigo
                    )
            return redirect("usuarios_lista")
    else:
        ensayos_actuales = LabEnsayo.objects.filter(laboratorio=laboratorio).select_related("ensayo")
        initial = [{
            "ensayo": le.ensayo.id,
            "codigo_ensayo": le.codigo_ensayo
        } for le in ensayos_actuales]
        formset = EditarEnsayosFormSet(initial=initial)

    return render(request, "editar_ensayos.html", {
        "laboratorio": laboratorio,
        "formset": formset
    })




@csrf_exempt
def login_view(request):
    """
    🚀 Vista de inicio de sesión que autentica al usuario y lo redirige según su grupo.
    ✅ Redirige a `index` si el usuario pertenece al grupo 'laboratorio'.
    ✅ Redirige a `usuarios_lista` si el usuario no pertenece a 'laboratorio'.
    
    Args:
        request (HttpRequest): La solicitud HTTP.

    Returns:
        HttpResponseRedirect: Redirige según el grupo del usuario.
    """
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)  # ✅ Autentica al usuario
            
            # 📌 Verifica el grupo del usuario y redirige en consecuencia
            if user.groups.filter(name="laboratorio").exists():
                return redirect("index")  # ✅ Redirige si pertenece al grupo "laboratorio"
            else:
                return redirect("usuarios_lista")  # ✅ Redirige si no pertenece al grupo
            
        else:
            return render(request, "login.html", {"error": "Usuario o contraseña incorrectos"})

    return render(request, "login.html")


@csrf_exempt
def convertir_aie(request):
    if request.method == 'POST':
        if 'archivo' not in request.FILES:
            return JsonResponse({'error': 'No se proporcionó un archivo válido'}, status=400)

        try:
            archivo = request.FILES['archivo']
            excel_file = pd.read_excel(archivo, engine="openpyxl")

            # Registrar columnas para depuración
            print("📌 Columnas detectadas en el servidor:", excel_file.columns)
            print("📌 Primeras filas del archivo:", excel_file.head())

            # Registrar en la base, pero no cortar si falla
            try:
                registrar_excel_aie(request, excel_file)
                print("✅ Se registraron los datos en la base")
            except Exception as e:
                print(f"⚠️ Error al registrar en la base: {str(e)}")
                traceback.print_exc()


            # Convertir columnas específicas a tipo string
            columns_to_str = [
                "Nro Informe", "RENSPA", "CUIT Funcionario", "Identificacion Muestra", "Identificacion Interna Laboratorio", "Sexo", 
                "Antigeno/Kit", "Marca Antigeno/Kit", "Lote", "Estampilla"
            ]
            for column in columns_to_str:
                excel_file[column] = excel_file[column].astype(str)

            # Aplicar el formato de fecha a las columnas correspondientes
            columnas_de_fecha = ["Fecha de Toma", "Fecha de Recepcion", "Fecha Inicio", "Fecha Fin", "Fecha Vencimiento Antigeno/Kit"]
            for columna in columnas_de_fecha:
                if columna in excel_file.columns:
                    excel_file[columna] = excel_file[columna].apply(formatear_fecha)

            # Agrupar datos generales sin afectar submuestras
            submuestras_dict= excel_file.groupby("Nro Informe").apply(procesar_submuestras).to_dict()
            excel_file_agrupado = agrupar_por_numero_informe_aie(excel_file)

            print("📌 Columnas después de agrupar:", list(excel_file.columns))
            print("📌 Primeras filas después de agrupar:", excel_file.head())

            
            # Procesar correctamente las submuestras
            print("📌 Antes de procesar submuestras")
            excel_file_agrupado["subMuestras"] = excel_file_agrupado["Nro Informe"].map(submuestras_dict)
            
            print("📌 Después de procesar submuestras")
            
            # ✅ Agregar depuración para verificar que las submuestras no están vacías
            print("📌 Submuestras después de procesar:", excel_file_agrupado["subMuestras"].head())


            # Convertir a JSON
            rubrosLab = rubros_lab(request)
            json_data = excel_file_agrupado.apply(lambda row: construir_json_aie(row, rubrosLab), axis=1).tolist()
            print("📌 JSON generado:", json_data[:2])  # Ver las primeras 2 estructuras

            # Guardar el JSON en un archivo temporal
            with NamedTemporaryFile(delete=False, suffix='.json') as temp_file:
                file_path = temp_file.name
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(json_data, f, ensure_ascii=False, indent=4)

            # Obtener la fecha actual y formatearla como dd_mm_yy
            fecha_actual = datetime.now().strftime("%d_%m_%y")
            nombre_archivo = f"AIE_{fecha_actual}.json"

            # Enviar el archivo como descarga con el nombre correcto
            response = HttpResponse(open(file_path, 'rb'), content_type='application/json')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            os.unlink(file_path)
            return response

        except Exception as e:
            import traceback
            error_msg = traceback.format_exc()
            print(f"❌ Error en la conversión:\n{error_msg}")
            return JsonResponse({'error': f'Ocurrió un error en la conversión: {str(e)}'}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)




csrf_exempt
def convertir_bru(request):
    if request.method == 'POST':
        if 'archivo' not in request.FILES:
            return JsonResponse({'error': 'No se proporcionó un archivo válido'}, status=400)

        try:
            archivo = request.FILES['archivo']
            excel_file = pd.read_excel(archivo, engine="openpyxl")

            try:
                registrar_excel_aie(request, excel_file)
                print("✅ Se registraron los datos en la base")
            except Exception as e:
                print(f"⚠️ Error al registrar en la base: {str(e)}")
                traceback.print_exc()

            # Convertir columnas específicas a tipo string
            columns_to_str = [
                "Nro Informe", "RENSPA", "CUIT Funcionario", "Especie", "Identificacion Muestra", "Identificacion Interna Laboratorio", "Sexo",
                "Antigeno/Kit", "Marca Antigeno/Kit", "Lote", "Estampilla"
            ]
            for column in columns_to_str:
                excel_file[column] = excel_file[column].astype(str)

            # Aplicar el formato de fecha a las columnas correspondientes
            columnas_de_fecha = ["Fecha de Toma", "Fecha de Recepcion", "Fecha Inicio", "Fecha Fin", "Fecha Vencimiento Antigeno/Kit"]
            for columna in columnas_de_fecha:
                if columna in excel_file.columns:
                    excel_file[columna] = excel_file[columna].apply(formatear_fecha)

            # Agrupar solo las columnas generales por número de informe
            plantilla_agrupada = agrupar_por_numero_informe_bru(excel_file)

            # Convertir columnas numéricas explícitamente y mantener NaN sin modificar
            numeric_columns = ["Motivo", "SubMotivo", "Cantidad Muestras", "Unidad de Medida", "Codigo DT"]
            for col in numeric_columns:
                plantilla_agrupada[col] = pd.to_numeric(plantilla_agrupada[col], errors='coerce')

            # Generar el análisis y submuestras correctamente agrupadas
            rubrosLab = rubros_lab(request)
            analisis_dict = generar_analisis_bru(excel_file, rubrosLab)

            # Asignar correctamente cada informe con sus análisis y submuestras
            plantilla_agrupada["analisis"] = plantilla_agrupada["Nro Informe"].map(lambda x: analisis_dict.get(x, []))

            # Construcción del JSON final
            codigoLaboratorio = rubrosLab["Nro_Laboratorio"].iloc[0] if not rubrosLab.empty else None
            json_data = plantilla_agrupada.apply(lambda row: construir_json_bru(row, codigoLaboratorio), axis=1).tolist()

            # Guardar el JSON en un archivo temporal
            with NamedTemporaryFile(delete=False, suffix='.json') as temp_file:
                file_path = temp_file.name
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(json_data, f, ensure_ascii=False, indent=4, default=convertir_tipo)

            # Obtener la fecha actual y formatearla como dd_mm_yy
            fecha_actual = datetime.now().strftime("%d_%m_%y")
            nombre_archivo = f"Brucelosis_{fecha_actual}.json"

            # Enviar el archivo como descarga con el nombre correcto
            response = HttpResponse(open(file_path, 'rb'), content_type='application/json')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            os.unlink(file_path)
            return response

        except Exception as e:
            import traceback
            error_msg = traceback.format_exc()
            print(f"❌ Error en la conversión:\n{error_msg}")
            return JsonResponse({'error': f'Ocurrió un error en la conversión: {str(e)}'}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)



@csrf_exempt
def convertir_Triqui(request):
    if request.method == 'POST':
        if 'archivo' not in request.FILES:
            return JsonResponse({'error': 'No se proporcionó un archivo válido'}, status=400)

        try:
            archivo = request.FILES['archivo']
            excel_file = pd.read_excel(archivo)

            # Convertir columnas específicas a tipo string
            columns_to_str = [
                "Nro Informe", "Nro Laboratorio", "Establecimiento", "Nro Documento",
                "CUIT Funcionario", "Identificacion Muestra"
            ]
            for column in columns_to_str:
                excel_file[column] = excel_file[column].astype(str)

            # Aplicar formato de fecha
            columnas_de_fecha = ["Fecha de Toma", "Fecha de Recepcion", "Fecha Inicio", "Fecha conclusion"]
            for columna in columnas_de_fecha:
                if columna in excel_file.columns:
                    excel_file[columna] = excel_file[columna].apply(formatear_fecha)

            # Agrupar datos generales sin afectar submuestras
            excel_file_agrupado = agrupar_por_informe_Triqui(excel_file)
            print(f"excel_file_agrupado_1{excel_file_agrupado.columns}")
            submuestras_dict = excel_file.groupby(["Nro Informe"]).apply(procesar_submuestras_triqui).to_dict()
            print(f"submuestras_dict_triqui:{submuestras_dict}")
            # Vincular submuestras a cada informe
            
            excel_file_agrupado["subMuestras"] = excel_file_agrupado["Nro Informe"].map(submuestras_dict)
            
    
            # Generar lista JSON
            json_data = excel_file_agrupado.apply(construir_json_triqui, axis=1).tolist()

            # Guardar en archivo temporal
            with NamedTemporaryFile(delete=False, suffix='.json') as temp_file:
                file_path = temp_file.name
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(json_data, f, ensure_ascii=False, indent=4, default=convertir_tipo)

            # Generar nombre del archivo final
            fecha_actual = datetime.now().strftime("%d_%m_%y")
            nombre_archivo = f"Triquinelosis_{fecha_actual}.json"

            # Preparar la respuesta de descarga
            response = HttpResponse(open(file_path, 'rb'), content_type='application/json')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

            # Eliminar archivo temporal una vez enviado
            os.unlink(file_path)
            return response

        except Exception as e:
            return JsonResponse({'error': f'Ocurrió un error en la conversión: {str(e)}'}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)


# NUEVA VERSION DE CONVERTIR TRIQUI - PROCESA EL EXCEL RESUMIDO Y LO CONVIERTE A JSON

@csrf_exempt
def convertir_Triqui_v2(request):# me esta faltando agregar una funcion que me traiga desde la BD el Nro lab, establecimiento y Còdigo DT
    if request.method == 'POST':
        if 'archivo' not in request.FILES:
            return JsonResponse({'error': 'No se proporcionó un archivo válido'}, status=400)

        try:
            archivo = request.FILES['archivo']
            excel_file = pd.read_excel(archivo)

            # Convertir columnas específicas a tipo string
            columns_to_str = [
                "Nro Informe", "Nro Autorización", "Nro Tropa", "Establecimiento",
                "CUIT Funcionario", "Identificacion Muestra", "Conclusion Protocolo"
            ]
            for column in columns_to_str:
                excel_file[column] = excel_file[column].astype(str)

            # Aplicar formato de fecha
            columnas_de_fecha = ["Fecha de Toma"]
            for columna in columnas_de_fecha:
                if columna in excel_file.columns:
                    excel_file[columna] = excel_file[columna].apply(formatear_fecha)

            # Agrupar datos generales sin afectar submuestras
            excel_file_agrupado = agrupar_informe_Triqui(excel_file)
            print(f"excel_file_agrupado_1{excel_file_agrupado.columns}")
            submuestras_dict = excel_file.groupby(["Nro Informe"]).apply(procesar_submuestras_triqui).to_dict()
            print(f"submuestras_dict_triqui:{submuestras_dict}")
            # Vincular submuestras a cada informe
            
            excel_file_agrupado["subMuestras"] = excel_file_agrupado["Nro Informe"].map(submuestras_dict)
            
    
            # Generar lista JSON
            json_data = excel_file_agrupado.apply(construir_json_triqui, axis=1).tolist()

            # Guardar en archivo temporal
            with NamedTemporaryFile(delete=False, suffix='.json') as temp_file:
                file_path = temp_file.name
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(json_data, f, ensure_ascii=False, indent=4, default=convertir_tipo)

            # Generar nombre del archivo final
            fecha_actual = datetime.now().strftime("%d_%m_%y")
            nombre_archivo = f"Triquinelosis_{fecha_actual}.json"

            # Preparar la respuesta de descarga
            response = HttpResponse(open(file_path, 'rb'), content_type='application/json')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

            # Eliminar archivo temporal una vez enviado
            os.unlink(file_path)
            return response

        except Exception as e:
            return JsonResponse({'error': f'Ocurrió un error en la conversión: {str(e)}'}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)






























    # Convierte el PDF del acta digital de Brucelosis a un excel para completar.

@csrf_exempt
def convertir_Acta_bru_old(request):
    # --- Consulta de códigos de rubro ---
    codigo_rubro_bru = codigoRubro(request, 2) # Poner el id del analito en la base de datos.
    codigo_rubro_bru = codigo_rubro_bru[['codigo_rubro', 'tecnica']]
    print(f"Rubro: {codigo_rubro_bru}")

    tecnicas = codigo_rubro_bru['tecnica'].unique()
    print(f"Tecnica laboratorio {tecnicas}")

    if len(tecnicas) > 1:
        tamiz_df = codigo_rubro_bru.loc[
            codigo_rubro_bru['tecnica'].str.contains("BPAT", case=False, na=False),
            'codigo_rubro'
        ]
        codigo_rubro = int(tamiz_df.iloc[0]) if not tamiz_df.empty else None
    elif len(tecnicas) == 1:
        unico_df = codigo_rubro_bru['codigo_rubro']
        codigo_rubro = int(unico_df.iloc[0]) if not unico_df.empty else None
    else:
        codigo_rubro = None
    print(f"Codigo rubro:{codigo_rubro}")
    # --- Procesamiento del POST ---
    if request.method == 'POST':
        try:
            # Definimos columnas del DataFrame final
            columnas = [
                "Nro Informe","Nro Acta","RENSPA","Motivo","SubMotivo","CUIT Funcionario","Fecha de Toma","Fecha de Recepcion","Especie","Cantidad Muestras",
                "Rubro","Fecha Inicio","Fecha Fin",
                "Resultado Letra","Identificacion Muestra","Identificacion Interna Laboratorio", "Tipo Identificacion","Observacion Muestra",
                "Categoria","Sexo","Antigeno/Kit","Marca Antigeno/Kit","Lote",
                "Fecha Vencimiento Antigeno/Kit","Estampilla","Codigo DT",
                "Observacion del Protocolo","Conclusion Protocolo"
            ]
            df_acta = pd.DataFrame(columns=columnas)

            #  Bucle para procesar hasta 4 actas
            for i in range(1, 5):
                acta = request.FILES.get(f"acta{i}")
                nro_informe = request.POST.get(f"informe{i}")
                fechaRecepcion= request.POST.get(f"fechaRecepcion5{i}")
                fechaInicio= request.POST.get(f"fechaInicio{i}")
                fechaFin= request.POST.get(f"fechaFin{i}")
                fechaVencimiento=request.POST.get(f"fechaVencimiento{i}")
                resultadoLetra= request.POST.get(f"resultadoLetra{i}")
                antigeno= request.POST.get(f"antigeno{i}")
                marcaAntigeno= request.POST.get(f"marcaAntigeno{i}")
                lote= request.POST.get(f"lote{i}")
                estampilla= request.POST.get(f"estampilla{i}")
                codigoDT= request.POST.get(f"codigoDT{i}")

                if acta and nro_informe:
                    df_tablas = extraer_tablas2(acta)
                    datospdf= extraer_datos_pdf (acta)
                    cant_muestras= int(len(df_tablas))
                    numero_acta = datospdf["numeroActa"]
                    cuitDeFuncionario = datospdf["cuitDeFuncionario"]
                    RENSPA = datospdf["RENSPA"]
                    motivo= datospdf["Motivo"]
                    submotivo= datospdf["SubMotivo"]
                    fechaToma= datospdf["FechaToma"]
                    especie=datospdf["Especie"]

                    print(f"Acta {i}: {len(df_tablas)} identificaciones")
                    print(f"Número de acta encontrado: {numero_acta}")
                    print(f"CUIT del Funcionario: {cuitDeFuncionario}")
                    print(f"Nro. Oficial Senasa: {RENSPA}")
                    print(f"Motivo: {motivo}")
                    print(f"submotivo: {submotivo}")
                    print(f"FechaToma: {fechaToma}")
                    print(f"Especie: {especie}")
                    

                    # Expandimos cada identificación en una fila
                    df_temp = pd.DataFrame({
                        "Nro Informe": [nro_informe] * len(df_tablas),
                        "Nro Acta": [numero_acta] * len(df_tablas),
                        "RENSPA": [RENSPA] * len(df_tablas),
                        "Motivo":[motivo]* len(df_tablas),
                        "SubMotivo":[submotivo]* len(df_tablas),
                        "CUIT Funcionario":[cuitDeFuncionario]* len(df_tablas),
                        "Fecha de Toma":[fechaToma] * len(df_tablas),
                        "Fecha de Recepcion": [fechaRecepcion] * len(df_tablas),
                        "Especie": [especie] * len(df_tablas),
                        "Cantidad Muestras": [cant_muestras]* len(df_tablas),
                        "Rubro": [codigo_rubro] * len(df_tablas),
                        "Fecha Inicio": [fechaInicio] * len(df_tablas),
                        "Fecha Fin": [fechaFin] * len(df_tablas), 
                        "Resultado Letra": [resultadoLetra] * len(df_tablas),
                        "Identificacion Muestra": df_tablas["Identificacion"].astype(str),
                        "Identificacion Interna Laboratorio": df_tablas["Nro_Tubo"],                        
                        "Tipo Identificacion": [1] * len(df_tablas),
                        "Observacion Muestra": None,
                        "Categoria": df_tablas["Categoria"],
                        "Sexo": df_tablas["Sexo"],
                        "Antigeno/Kit": [antigeno] * len(df_tablas),
                        "Marca Antigeno/Kit": [marcaAntigeno] * len(df_tablas),
                        "Lote": [lote] * len(df_tablas),
                        "Fecha Vencimiento Antigeno/Kit": [fechaVencimiento] * len(df_tablas),  # fecha
                        "Estampilla": [estampilla] * len(df_tablas),
                        "Codigo DT": [codigoDT] * len(df_tablas),
                        "Observacion del Protocolo": None,
                        "Conclusion Protocolo": None,
                    })

                    # Concatenamos al DataFrame final
                    df_acta = pd.concat([df_acta, df_temp], ignore_index=True)
            # --- Conversión de tipos antes de exportar ---
            columnas_de_fecha = [
                "Fecha de Toma","Fecha de Recepcion","Fecha Inicio","Fecha Fin","Fecha Vencimiento Antigeno/Kit"
            ]
            for columna in columnas_de_fecha:
                if columna in df_acta.columns:
                    df_acta[columna] = df_acta[columna].apply(formatear_fecha)

            columnas_enteros = ["Nro Informe","Nro Acta","Cantidad Muestras","Rubro"]
            for columna in columnas_enteros:
                if columna in df_acta.columns:
                    df_acta[columna] = pd.to_numeric(df_acta[columna], errors="coerce").astype("Int64")

            # --- Generar Excel ---
            fecha_actual = datetime.now().strftime("%Y-%m-%d")
            nombre_archivo = f"Acta_Brucelosis_{fecha_actual}.xlsx"

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

            with pd.ExcelWriter(response, engine='openpyxl') as writer:
                df_acta.to_excel(writer, sheet_name="Acta Digital", index=False, header=True)

            return response

        except Exception as e:
            return JsonResponse({'error': f'Ocurrió un error en la conversión: {str(e)}'}, status=500)

    return JsonResponse({'error': 'No se pudo generar el archivo excel'}, status=405)


#Convierte el pdf del acta digital AIE y descarga un excel.
@csrf_exempt
def convertir_Acta_AIE(request):
    # --- Consulta de códigos de rubro ---
    codigo_rubro_aie = codigoRubro(request, 1)  # Poner el id del analito en la base de datos.
    codigo_rubro_aie = codigo_rubro_aie[['codigo_rubro']]
    codigo_rubro_aie = int(codigo_rubro_aie.iloc[0]) if not codigo_rubro_aie.empty else None
    print(f"Rubro: {codigo_rubro_aie}")

    # --- Procesamiento del POST ---
    if request.method == 'POST':
        try:
            # Definimos columnas del DataFrame final
            columnas = [
                "Nro Informe","Nro Acta","RENSPA","Motivo",
                "SubMotivo","CUIT Funcionario","Fecha de Toma","Fecha de Recepcion",
                "Especie","Cantidad Muestras","Rubro","Fecha Inicio","Fecha Fin",
                "Resultado Letra","Identificacion Muestra","Identificacion Interna Laboratorio", "Tipo Identificacion","Observacion Muestra",
                "Categoria","Sexo","Antigeno/Kit","Marca Antigeno/Kit","Lote",
                "Fecha Vencimiento Antigeno/Kit","Estampilla","Codigo DT",
                "Observacion del Protocolo","Conclusion Protocolo"
            ]
            df_acta = pd.DataFrame(columns=columnas)

            #  Bucle para procesar hasta 4 actas
            for i in range(1, 5):
                acta = request.FILES.get(f"acta{i}")
                nro_informe = request.POST.get(f"informe{i}")
                fechaRecepcion= request.POST.get(f"fechaRecepcion5{i}")
                fechaInicio= request.POST.get(f"fechaInicio{i}")
                fechaFin= request.POST.get(f"fechaFin{i}")
                fechaVencimiento=request.POST.get(f"fechaVencimiento{i}")
                resultadoLetra= request.POST.get(f"resultadoLetra{i}")
                antigeno= request.POST.get(f"antigeno{i}")
                marcaAntigeno= request.POST.get(f"marcaAntigeno{i}")
                lote= request.POST.get(f"lote{i}")
                estampilla= request.POST.get(f"estampilla{i}")
                codigoDT= request.POST.get(f"codigoDT{i}")
                conclusion= request.POST.get(f"conclusion{i}")

                if acta and nro_informe:
                    processor = PDFProcessor(acta)
                    if not processor.es_valido():
                        logger.warning(f"Acta {nro_informe}: PDF sin texto legible (posible imagen sin OCR).")
                        return JsonResponse({'error': 'El PDF no contiene texto legible (posible imagen sin OCR).'}, status=400)

                    df_tablas = processor.extraer_tablas()
                    datospdf = processor.extraer_datos_cabecera()
                    processor.cerrar()
                    cant_muestras = int(len(df_tablas))
                    numero_acta = datospdf["numeroActa"]
                    cuitDeFuncionario = datospdf["cuitDeFuncionario"]
                    RENSPA = datospdf["RENSPA"]
                    motivo = datospdf["Motivo"]
                    submotivo = datospdf["SubMotivo"]
                    fechaToma = datospdf["FechaToma"]
                    especie = datospdf["Especie"]

                    print(f"Acta {i}: {len(df_tablas)} identificaciones")
                    print(f"Número de acta encontrado: {numero_acta}")
                    print(f"CUIT del Funcionario: {cuitDeFuncionario}")
                    print(f"Nro. Oficial Senasa: {RENSPA}")
                    print(f"Motivo: {motivo}")
                    print(f"submotivo: {submotivo}")
                    print(f"FechaToma: {fechaToma}")
                    print(f"Especie: {especie}")

                    # Expandimos cada identificación en una fila
                    df_temp = pd.DataFrame({
                        "Nro Informe": [nro_informe] * len(df_tablas),
                        "Nro Acta": [numero_acta] * len(df_tablas),
                        "RENSPA": [RENSPA] * len(df_tablas),
                        "Motivo": [motivo] * len(df_tablas),
                        "SubMotivo": [submotivo] * len(df_tablas),
                        "CUIT Funcionario": [cuitDeFuncionario] * len(df_tablas),
                        "Fecha de Toma": [fechaToma] * len(df_tablas),
                        "Fecha de Recepcion": [fechaRecepcion] * len(df_tablas),
                        "Especie": [especie] * len(df_tablas),
                        "Cantidad Muestras": [cant_muestras] * len(df_tablas),
                        "Rubro": [codigo_rubro_aie] * len(df_tablas),
                        "Fecha Inicio": [fechaInicio] * len(df_tablas),
                        "Fecha Fin": [fechaFin] * len(df_tablas),
                        "Resultado Letra": [resultadoLetra] * len(df_tablas),
                        "Identificacion Muestra": df_tablas["Identificacion"].astype(str),
                        "Identificacion Interna Laboratorio": df_tablas["Nro_Tubo"],
                        "Tipo Identificacion": 12,
                        "Observacion Muestra": None,
                        "Categoria": df_tablas["Categoria"],
                        "Sexo": df_tablas["Sexo"],
                        "Antigeno/Kit": [antigeno] * len(df_tablas),
                        "Marca Antigeno/Kit": [marcaAntigeno] * len(df_tablas),
                        "Lote": [lote] * len(df_tablas),
                        "Fecha Vencimiento Antigeno/Kit": [fechaVencimiento] * len(df_tablas),
                        "Estampilla": [estampilla] * len(df_tablas),
                        "Codigo DT": [codigoDT] * len(df_tablas),
                        "Observacion del Protocolo": None,
                        "Conclusion Protocolo": [conclusion] * len(df_tablas),
                    })

                    # Concatenamos al DataFrame final
                    df_acta = pd.concat([df_acta, df_temp], ignore_index=True)

            # --- Conversión de tipos antes de exportar ---
            columnas_de_fecha = [
                "Fecha de Toma","Fecha de Recepcion","Fecha Inicio","Fecha Fin","Fecha Vencimiento Antigeno/Kit"
            ]
            for columna in columnas_de_fecha:
                if columna in df_acta.columns:
                    df_acta[columna] = df_acta[columna].apply(formatear_fecha)

            columnas_enteros = ["Cantidad Muestras","Rubro"]
            for columna in columnas_enteros:
                if columna in df_acta.columns:
                    df_acta[columna] = pd.to_numeric(df_acta[columna], errors="coerce").astype("Int64")

            # --- Generar Excel ---
            fecha_actual = datetime.now().strftime("%d-%m-%Y")
            nombre_archivo = f"Acta_AIE_{fecha_actual}.xlsx"

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

            # Exportar con formato de fecha real
            with pd.ExcelWriter(response, engine='openpyxl', datetime_format='DD/MM/YYYY') as writer:
                df_acta.to_excel(writer, sheet_name="Acta Digital", index=False, header=True)

            return response

        except Exception as e:
            return JsonResponse({'error': f'Ocurrió un error en la conversión: {str(e)}'}, status=500)

    return JsonResponse({'error': 'No se pudo generar el archivo excel'}, status=405)

@csrf_exempt
def convertir_Acta_Auj(request):
    # --- Consulta de códigos de rubro ---
    codigo_rubro_auj = codigoRubro(request, 5)  # Poner el id del analito en la base de datos.
    codigo_rubro_auj = codigo_rubro_auj[['codigo_rubro']]
    codigo_rubro_auj = int(codigo_rubro_auj.iloc[0]) if not codigo_rubro_auj.empty else None
    print(f"Rubro: {codigo_rubro_auj}")

    # --- Procesamiento del POST ---
    if request.method == 'POST':
        try:
            # Definimos columnas del DataFrame final
            columnas = [
                "Nro Informe","Nro Acta","RENSPA","Motivo",
                "SubMotivo","CUIT Funcionario","Fecha de Toma","Fecha de Recepcion",
                "Especie","Cantidad Muestras","Rubro","Fecha Inicio","Fecha Fin",
                "Resultado Letra","Identificacion Muestra","Identificacion Interna Laboratorio", "Tipo Identificacion","Observacion Muestra",
                "Categoria","Sexo","Antigeno/Kit","Marca Antigeno/Kit","Lote",
                "Fecha Vencimiento Antigeno/Kit","Estampilla","Codigo DT",
                "Observacion del Protocolo","Conclusion Protocolo"
            ]
            df_acta = pd.DataFrame(columns=columnas)

            #  Bucle para procesar hasta 4 actas
            for i in range(1, 5):
                acta = request.FILES.get(f"acta{i}")
                nro_informe = request.POST.get(f"informe{i}")
                fechaRecepcion= request.POST.get(f"fechaRecepcion5{i}")
                fechaInicio= request.POST.get(f"fechaInicio{i}")
                fechaFin= request.POST.get(f"fechaFin{i}")
                fechaVencimiento=request.POST.get(f"fechaVencimiento{i}")
                resultadoLetra= request.POST.get(f"resultadoLetra{i}")
                antigeno= request.POST.get(f"antigeno{i}")
                marcaAntigeno= request.POST.get(f"marcaAntigeno{i}")
                lote= request.POST.get(f"lote{i}")
                estampilla= request.POST.get(f"estampilla{i}")
                codigoDT= request.POST.get(f"codigoDT{i}")
                conclusion= request.POST.get(f"conclusion{i}")

                if acta and nro_informe:
                    df_tablas = extraer_tablas2(acta)
                    datospdf = extraer_datos_pdf(acta)
                    cant_muestras = int(len(df_tablas))
                    numero_acta = datospdf["numeroActa"]
                    cuitDeFuncionario = datospdf["cuitDeFuncionario"]
                    RENSPA = datospdf["RENSPA"]
                    motivo = datospdf["Motivo"] #importar las tablas de motivos y sumotivos a la base de datos para que los pueda buscar
                    submotivo = datospdf["SubMotivo"]
                    fechaToma = datospdf["FechaToma"]
                    especie = datospdf["Especie"] # importar las especies de porcinos a la base de datos

                    print(f"Acta {i}: {len(df_tablas)} identificaciones")
                    print(f"Número de acta encontrado: {numero_acta}")
                    print(f"CUIT del Funcionario: {cuitDeFuncionario}")
                    print(f"Nro. Oficial Senasa: {RENSPA}")
                    print(f"Motivo: {motivo}")
                    print(f"submotivo: {submotivo}")
                    print(f"FechaToma: {fechaToma}")
                    print(f"Especie: {especie}")

                    # Expandimos cada identificación en una fila
                    df_temp = pd.DataFrame({
                        "Nro Informe": [nro_informe] * len(df_tablas),
                        "Nro Acta": [numero_acta] * len(df_tablas),
                        "RENSPA": [RENSPA] * len(df_tablas),
                        "Motivo": [motivo] * len(df_tablas),
                        "SubMotivo": [submotivo] * len(df_tablas),
                        "CUIT Funcionario": [cuitDeFuncionario] * len(df_tablas),
                        "Fecha de Toma": [fechaToma] * len(df_tablas),
                        "Fecha de Recepcion": [fechaRecepcion] * len(df_tablas),
                        "Especie": [especie] * len(df_tablas),
                        "Cantidad Muestras": [cant_muestras] * len(df_tablas),
                        "Rubro": [codigo_rubro_auj] * len(df_tablas),
                        "Fecha Inicio": [fechaInicio] * len(df_tablas),
                        "Fecha Fin": [fechaFin] * len(df_tablas),
                        "Resultado Letra": [resultadoLetra] * len(df_tablas),
                        "Identificacion Muestra": df_tablas["Identificacion"].astype(str),
                        "Identificacion Interna Laboratorio": df_tablas["Nro_Tubo"],
                        "Tipo Identificacion": 12,
                        "Observacion Muestra": None,
                        "Categoria": df_tablas["Categoria"],
                        "Sexo": df_tablas["Sexo"],
                        "Antigeno/Kit": [antigeno] * len(df_tablas),
                        "Marca Antigeno/Kit": [marcaAntigeno] * len(df_tablas),
                        "Lote": [lote] * len(df_tablas),
                        "Fecha Vencimiento Antigeno/Kit": [fechaVencimiento] * len(df_tablas),
                        "Estampilla": [estampilla] * len(df_tablas),
                        "Codigo DT": [codigoDT] * len(df_tablas),
                        "Observacion del Protocolo": None,
                        "Conclusion Protocolo": conclusion * len(df_tablas),
                    })

                    # Concatenamos al DataFrame final
                    df_acta = pd.concat([df_acta, df_temp], ignore_index=True)

            # --- Conversión de tipos antes de exportar ---
            columnas_de_fecha = [
                "Fecha de Toma","Fecha de Recepcion","Fecha Inicio","Fecha Fin","Fecha Vencimiento Antigeno/Kit"
            ]
            for columna in columnas_de_fecha:
                if columna in df_acta.columns:
                    df_acta[columna] = df_acta[columna].apply(formatear_fecha)

            columnas_enteros = ["Nro Informe","Nro Acta","Cantidad Muestras","Rubro"]
            for columna in columnas_enteros:
                if columna in df_acta.columns:
                    df_acta[columna] = pd.to_numeric(df_acta[columna], errors="coerce").astype("Int64")

            # --- Generar Excel ---
            fecha_actual = datetime.now().strftime("%d-%m-%Y")
            nombre_archivo = f"Acta_AUJESZKY_{fecha_actual}.xlsx"

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

            # Exportar con formato de fecha real
            with pd.ExcelWriter(response, engine='openpyxl', datetime_format='DD/MM/YYYY') as writer:
                df_acta.to_excel(writer, sheet_name="Acta Digital", index=False, header=True)

            return response

        except Exception as e:
            return JsonResponse({'error': f'Ocurrió un error en la conversión: {str(e)}'}, status=500)

    return JsonResponse({'error': 'No se pudo generar el archivo excel'}, status=405)

# Convertir el EXCEL del acta digital a JSON

csrf_exempt
def crear_json_bru_digital(request):
    if request.method == 'POST':
        if 'archivo' not in request.FILES:
            return JsonResponse({'error': 'No se proporcionó un archivo válido'}, status=400)
      

        try:
            archivo = request.FILES['archivo']
            ext = os.path.splitext(archivo.name)[1].lower()
            if ext == ".xls":
                excel_file = pd.read_excel(archivo, engine="xlrd")
            else:
                excel_file = pd.read_excel(archivo, engine="openpyxl")

            
            # Convertir columnas específicas a tipo string
            columns_to_str = [
                "Nro Informe","Nro Acta", "RENSPA", "CUIT Funcionario", "Especie", "Identificacion Muestra", "Identificacion Interna Laboratorio", "Sexo",
                "Antigeno/Kit", "Marca Antigeno/Kit", "Lote", "Estampilla"
            ]
            for column in columns_to_str:
                excel_file[column] = excel_file[column].astype(str)

            # Aplicar el formato de fecha a las columnas correspondientes
            columnas_de_fecha = ["Fecha de Toma", "Fecha de Recepcion", "Fecha Inicio", "Fecha Fin", "Fecha Vencimiento Antigeno/Kit"]
            for columna in columnas_de_fecha:
                if columna in excel_file.columns:
                    excel_file[columna] = excel_file[columna].apply(formatear_fecha)
            print(f"fecha corregida: {excel_file[columna]}")

            # Agrupar solo las columnas generales por número de informe
            plantilla_agrupada = agrupar_columnas_bru(excel_file)

            # Convertir columnas numéricas explícitamente y mantener NaN sin modificar
            numeric_columns = ["Motivo", "SubMotivo","Cantidad Muestras","Codigo DT"]
            for col in numeric_columns:
                plantilla_agrupada[col] = pd.to_numeric(plantilla_agrupada[col], errors='coerce')

            # Generar el análisis y submuestras correctamente agrupadas
            rubrosLab = rubros_lab(request)
            analisis_dict = analisis_bru_digital(excel_file)

            # Asignar correctamente cada informe con sus análisis y submuestras
            plantilla_agrupada["analisis"] = plantilla_agrupada["Nro Informe"].map(lambda x: analisis_dict.get(x, []))

            # Construcción del JSON final
            codigoLaboratorio = rubrosLab["Nro_Laboratorio"].iloc[0] if not rubrosLab.empty else None
            json_data = plantilla_agrupada.apply(lambda row: json_bru(row, codigoLaboratorio), axis=1).tolist()

            # Guardar el JSON en un archivo temporal
            with NamedTemporaryFile(delete=False, suffix='.json') as temp_file:
                file_path = temp_file.name
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(json_data, f, ensure_ascii=False, indent=4, default=convertir_tipo)

            # Obtener la fecha actual y formatearla como dd_mm_yy
            fecha_actual = datetime.now().strftime("%d_%m_%y")
            nombre_archivo = f"Brucelosis_Acta_Digital_{fecha_actual}.json"

            # Enviar el archivo como descarga con el nombre correcto
            response = HttpResponse(open(file_path, 'rb'), content_type='application/json')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            os.unlink(file_path)
            return response

        except Exception as e:
            import traceback
            error_msg = traceback.format_exc()
            print(f"❌ Error en la conversión:\n{error_msg}")
            return JsonResponse({'error': f'Ocurrió un error en la conversión: {str(e)}'}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)





csrf_exempt
def crear_json_AIE_digital(request):
    if request.method == 'POST':
        if 'archivo' not in request.FILES:
            return JsonResponse({'error': 'No se proporcionó un archivo válido'}, status=400)

        try:
            archivo = request.FILES['archivo']
            ext = os.path.splitext(archivo.name)[1].lower()
            if ext == ".xls":
                excel_file = pd.read_excel(archivo, engine="xlrd")
            else:
                excel_file = pd.read_excel(archivo, engine="openpyxl")

            # Convertir columnas específicas a tipo string
            columns_to_str = [
                "Nro Informe", "RENSPA", "CUIT Funcionario",
                 "Identificacion Muestra", "Identificacion Interna Laboratorio", "Sexo",
                   "Antigeno/Kit", "Marca Antigeno/Kit", "Lote", 
                   "Estampilla"
                   ]
            for column in columns_to_str:
                excel_file[column] = excel_file[column].astype(str)

            # Aplicar el formato de fecha a las columnas correspondientes
            columnas_de_fecha = ["Fecha de Toma", "Fecha de Recepcion",
                                 "Fecha Inicio", "Fecha Fin", "Fecha Vencimiento Antigeno/Kit"]
            for columna in columnas_de_fecha:
                if columna in excel_file.columns:
                    excel_file[columna] = excel_file[columna].apply(formatear_fecha)

            submuestras_dict= excel_file.groupby("Nro Informe").apply(submuestras).to_dict()
            plantilla_agrupada = agrupar_columnas_aie(excel_file)
           

            print("📌 Columnas después de agrupar:", list(excel_file.columns))
            print("📌 Primeras filas después de agrupar:", excel_file.head())
            # Agrupar solo las columnas generales por número de informe
            
            
            print("📌 Antes de procesar submuestras")
            plantilla_agrupada["subMuestras"] = plantilla_agrupada["Nro Informe"].map(submuestras_dict)
            print("📌 Después de procesar submuestras")
            print("📌 Submuestras después de procesar:", plantilla_agrupada["subMuestras"].head())
            # Generar el análisis y submuestras correctamente agrupadas
            rubrosLab = rubros_lab(request)
            
            # Construcción del JSON final
            codigoLaboratorio = rubrosLab["Nro_Laboratorio"].iloc[0] if not rubrosLab.empty else None
            json_data = plantilla_agrupada.apply(lambda row: json_aie_digital(row, codigoLaboratorio), axis=1).tolist()
            print("📌 JSON generado:", json_data[:2])  # Ver las primeras 2 estructuras
            # Guardar el JSON en un archivo temporal
            with NamedTemporaryFile(delete=False, suffix='.json') as temp_file:
                file_path = temp_file.name
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(json_data, f, ensure_ascii=False, indent=4, default=convertir_tipo)

            # Obtener la fecha actual y formatearla como dd_mm_yy
            fecha_actual = datetime.now().strftime("%d_%m_%y")
            nombre_archivo = f"AIE_Acta_Digital_{fecha_actual}.json"

            # Enviar el archivo como descarga con el nombre correcto
            response = HttpResponse(open(file_path, 'rb'), content_type='application/json')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            os.unlink(file_path)
            return response

        except Exception as e:
            import traceback
            error_msg = traceback.format_exc()
            print(f"❌ Error en la conversión:\n{error_msg}")
            return JsonResponse({'error': f'Ocurrió un error en la conversión: {str(e)}'}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)

#En esta funcion convierto el PDF del acta digital directamente a JSON

#En esta funcion convierto el PDF del acta digital directamente a JSON

@csrf_exempt
def actaBru_JSON(request):
    # --- Consulta de códigos de rubro ---
    codigo_rubro_bru = codigoRubro(request, 2)
    codigo_rubro_bru = codigo_rubro_bru[['codigo_rubro', 'tecnica']]
    tecnicas = codigo_rubro_bru['tecnica'].unique()

    if len(tecnicas) > 1:
        tamiz_df = codigo_rubro_bru.loc[
            codigo_rubro_bru['tecnica'].str.contains("BPAT", case=False, na=False),
            'codigo_rubro'
        ]
        confirmatoria_df = codigo_rubro_bru.loc[
            codigo_rubro_bru['tecnica'].str.contains("FPA|SAT Y 2-ME|FCT|I-ELISA", case=False, na=False),
            'codigo_rubro'
        ]
        if not tamiz_df.empty:
            codigo_rubro_tamiz = int(tamiz_df.iloc[0])
        if not confirmatoria_df.empty:
            codigo_rubro_confirmatoria = int(confirmatoria_df.iloc[0])
    elif len(tecnicas) == 1:
        unico_df = codigo_rubro_bru['codigo_rubro']
        codigo_rubro_tamiz = int(unico_df.iloc[0]) if not unico_df.empty else None


    # --- Procesamiento del POST ---
    if request.method == 'POST':
        try:
            columnas = [
                "Nro Informe","Nro Acta","RENSPA","Motivo","SubMotivo","CUIT Funcionario","Fecha de Toma","Fecha de Recepcion","Especie","Cantidad Muestras",
                "Rubro","Fecha Inicio","Fecha Fin",
                "Resultado Letra","Identificacion Muestra","Identificacion Interna Laboratorio", "Tipo Identificacion","Observacion Muestra",
                "Categoria","Sexo","Antigeno/Kit","Marca Antigeno/Kit","Lote",
                "Fecha Vencimiento Antigeno/Kit","Estampilla","Codigo DT",
                "Observacion del Protocolo","Conclusion Protocolo"
            ]
            df_acta = pd.DataFrame(columns=columnas)

            # Bucle para procesar hasta 5 actas
            for i in range(1, 5):
                acta = request.FILES.get(f"acta{i}")
                nro_informe = request.POST.get(f"informe{i}")
                fechaRecepcion = request.POST.get(f"fechaRecepcion{i}")
                fechaInicio = request.POST.get(f"fechaInicio{i}")
                fechaFin = request.POST.get(f"fechaFin{i}")
                fechaVencimiento = request.POST.get(f"fechaVencimiento{i}")
                resultadoLetra = request.POST.get(f"resultadoLetra{i}")
                antigeno = request.POST.get(f"antigeno{i}")
                marcaAntigeno = request.POST.get(f"marcaAntigeno{i}")
                lote = request.POST.get(f"lote{i}")
                estampilla = request.POST.get(f"estampilla{i}")
                codigoDT = request.POST.get(f"codigoDT{i}")
                conclusion = request.POST.get(f"conclusion{i}")

                if acta and nro_informe:
                    processor = PDFProcessor(acta)
                    if not processor.es_valido():
                        logger.warning(f"Acta {nro_informe}: PDF sin texto legible (posible imagen sin OCR).")
                        return JsonResponse({'error': 'El PDF no contiene texto legible (posible imagen sin OCR).'}, status=400)

                    df_tablas = processor.extraer_tablas()
                    datospdf = processor.extraer_datos_cabecera()
                    processor.cerrar()
                    cant_muestras = int(len(df_tablas))

                    numero_acta = datospdf["numeroActa"]
                    cuitDeFuncionario = datospdf["cuitDeFuncionario"]
                    RENSPA = datospdf["RENSPA"]
                    motivo = datospdf["Motivo"]
                    submotivo = datospdf["SubMotivo"]
                    fechaToma = datospdf["FechaToma"]
                    especie = datospdf["Especie"]

                    resultadoLetra = int(resultadoLetra) if resultadoLetra and resultadoLetra.isdigit() else None       
                    categoria = df_tablas["Categoria"].apply(lambda x: int(x) if pd.notna(x) and str(x).isdigit() else None)

             
                    df_temp = pd.DataFrame({
                        "Nro Informe": [nro_informe] * len(df_tablas),
                        "Nro Acta": [numero_acta] * len(df_tablas),
                        "RENSPA": [RENSPA] * len(df_tablas),
                        "Motivo": [motivo] * len(df_tablas),
                        "SubMotivo": [submotivo] * len(df_tablas),
                        "CUIT Funcionario": [cuitDeFuncionario] * len(df_tablas),
                        "Fecha de Toma": [fechaToma] * len(df_tablas),
                        "Fecha de Recepcion": [fechaRecepcion] * len(df_tablas),
                        "Especie": [especie] * len(df_tablas),
                        "Cantidad Muestras": [cant_muestras] * len(df_tablas),
                        "Rubro": [codigo_rubro_tamiz] * len(df_tablas),
                        "Fecha Inicio": [fechaInicio] * len(df_tablas),
                        "Fecha Fin": [fechaFin] * len(df_tablas),
                        "Resultado Letra": [resultadoLetra] * len(df_tablas),
                        "Identificacion Muestra": df_tablas["Identificacion"].astype(str),
                        "Identificacion Interna Laboratorio": df_tablas["Nro_Tubo"],
                        "Tipo Identificacion": [1] * len(df_tablas),
                        "Observacion Muestra": [None] * len(df_tablas),
                        "Categoria": categoria,
                        "Sexo": df_tablas["Sexo"],
                        "Antigeno/Kit": [antigeno] * len(df_tablas),
                        "Marca Antigeno/Kit": [marcaAntigeno] * len(df_tablas),
                        "Lote": [lote] * len(df_tablas),
                        "Fecha Vencimiento Antigeno/Kit": [fechaVencimiento] * len(df_tablas),
                        "Estampilla": [estampilla] * len(df_tablas),
                        "Codigo DT": [codigoDT] * len(df_tablas),
                        "Observacion del Protocolo": [None] * len(df_tablas),
                        "Conclusion Protocolo": [conclusion]*len(df_tablas),
                    })

                    df_acta = pd.concat([df_acta, df_temp], ignore_index=True)

                    # --- Bloque complementario (confirmatoria) ---
                    extra_info = request.POST.get(f"toggleComp{i}")
                    if extra_info == "on":
                        fechaInicioComp = request.POST.get(f"fechaInicioComp{i}")
                        fechaFinComp = request.POST.get(f"fechaFinComp{i}")
                        fechaVencimientoComp = request.POST.get(f"fechaVencimientoComp{i}")
                        resultadoLetraComp = request.POST.get(f"resultadoLetraComp{i}")
                        antigenoComp = request.POST.get(f"antigenoComp{i}")
                        marcaAntigenoComp = request.POST.get(f"marcaAntigenoComp{i}")
                        loteComp = request.POST.get(f"loteComp{i}")
                        estampillaComp = request.POST.get(f"estampillaComp{i}")

                        resultadoLetraComp = int(resultadoLetraComp) if resultadoLetraComp and resultadoLetraComp.isdigit() else None

                        if any([fechaInicioComp, fechaFinComp, resultadoLetraComp, antigenoComp]):
                            df_comp = pd.DataFrame({
                                "Nro Informe": [nro_informe] * len(df_tablas),
                                "Nro Acta": [numero_acta] * len(df_tablas),
                                "RENSPA": [RENSPA] * len(df_tablas),
                                "Motivo": [motivo] * len(df_tablas),
                                "SubMotivo": [submotivo] * len(df_tablas),
                                "CUIT Funcionario": [cuitDeFuncionario] * len(df_tablas),
                                "Fecha de Toma": [formatear_fecha(fechaToma)] * len(df_tablas),
                                "Fecha de Recepcion": [formatear_fecha(fechaRecepcion)] * len(df_tablas),
                                "Especie": [especie] * len(df_tablas),
                                "Cantidad Muestras": [cant_muestras] * len(df_tablas),
                                "Rubro": [codigo_rubro_confirmatoria] * len(df_tablas),
                                "Fecha Inicio": [formatear_fecha(fechaInicioComp)] * len(df_tablas),
                                "Fecha Fin": [formatear_fecha(fechaFinComp)] * len(df_tablas),
                                "Resultado Letra": [resultadoLetraComp] * len(df_tablas),
                                "Identificacion Muestra": df_tablas["Identificacion"].astype(str),
                                "Identificacion Interna Laboratorio": df_tablas["Nro_Tubo"],
                                "Tipo Identificacion": [1] * len(df_tablas),
                                "Observacion Muestra": None,
                                "Categoria": categoria,
                                "Sexo": df_tablas["Sexo"],
                                "Antigeno/Kit": [antigenoComp] * len(df_tablas),
                                "Marca Antigeno/Kit": [marcaAntigenoComp] * len(df_tablas),
                                "Lote": [loteComp] * len(df_tablas),
                                "Fecha Vencimiento Antigeno/Kit": [formatear_fecha(fechaVencimientoComp)] * len(df_tablas),
                                "Estampilla": [estampillaComp] * len(df_tablas),
                                "Codigo DT": [codigoDT] * len(df_tablas),
                                "Observacion del Protocolo": None,
                                "Conclusion Protocolo": [conclusion] * len(df_tablas),
                            })
                            df_acta = pd.concat([df_acta, df_comp], ignore_index=True)

            # --- Conversión de tipos ---
            columnas_de_fecha = ["Fecha de Toma","Fecha de Recepcion","Fecha Inicio","Fecha Fin","Fecha Vencimiento Antigeno/Kit"]
            for columna in columnas_de_fecha:
                if columna in df_acta.columns:
                    df_acta[columna] = df_acta[columna].apply(formatear_fecha)

            columnas_enteros = ["Cantidad Muestras","Rubro"]
            for columna in columnas_enteros:
                if columna in df_acta.columns:
                    df_acta[columna] = pd.to_numeric(df_acta[columna], errors="coerce").astype("Int64")

            # --- Procesar DataFrame y generar JSON ---
            json_data = nexo_acta_json_bru(df_acta, request)

            fecha_actual = datetime.now().strftime("%d_%m_%y")
            nombre_archivo = f"Brucelosis_Acta_Digital_{fecha_actual}.json"

            response = HttpResponse(
                json.dumps(json_data, ensure_ascii=False, indent=4, default=convertir_tipo),
                content_type='application/json'
            )
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            return response

        except Exception as e:
            import traceback
            error_msg = traceback.format_exc()
            print(f"❌ Error en la conversión:\n{error_msg}")
            return JsonResponse({'error': f'Ocurrió un error en la conversión: {str(e)}'}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@csrf_exempt
def actaAIE_JSON(request):
    # --- Consulta de códigos de rubro ---
    codigo_rubro_aie = codigoRubro(request, 1)  # id del analito AIE en la base de datos
    codigo_rubro_aie = codigo_rubro_aie[['codigo_rubro']]
    codigo_rubro = int(codigo_rubro_aie.iloc[0]) if not codigo_rubro_aie.empty else None

    # --- Procesamiento del POST ---
    if request.method == 'POST':
        try:
            columnas = [
                "Nro Informe","Nro Acta","RENSPA","Motivo","SubMotivo","CUIT Funcionario",
                "Fecha de Toma","Fecha de Recepcion","Especie","Cantidad Muestras",
                "Rubro","Fecha Inicio","Fecha Fin","Resultado Letra",
                "Identificacion Muestra","Identificacion Interna Laboratorio",
                "Tipo Identificacion","Observacion Muestra","Categoria","Sexo",
                "Antigeno/Kit","Marca Antigeno/Kit","Lote","Fecha Vencimiento Antigeno/Kit",
                "Estampilla","Codigo DT","Observacion del Protocolo","Conclusion Protocolo"
            ]
            df_acta = pd.DataFrame(columns=columnas)

            # Bucle para procesar hasta 4 actas
            for i in range(1, 5):
                acta = request.FILES.get(f"acta{i}")
                nro_informe = request.POST.get(f"informe{i}")
                fechaRecepcion = request.POST.get(f"fechaRecepcion{i}")
                fechaInicio = request.POST.get(f"fechaInicio{i}")
                fechaFin = request.POST.get(f"fechaFin{i}")
                fechaVencimiento = request.POST.get(f"fechaVencimiento{i}")
                resultadoLetra = request.POST.get(f"resultadoLetra{i}")
                antigeno = request.POST.get(f"antigeno{i}")
                marcaAntigeno = request.POST.get(f"marcaAntigeno{i}")
                lote = request.POST.get(f"lote{i}")
                estampilla = request.POST.get(f"estampilla{i}")
                codigoDT = request.POST.get(f"codigoDT{i}")
                conclusion= request.POST.get(f"conclusion{i}")

                if acta and nro_informe:
                    processor = PDFProcessor(acta)
                    if not processor.es_valido():
                        logger.warning(f"Acta {nro_informe}: PDF sin texto legible (posible imagen sin OCR).")
                        return JsonResponse({'error': 'El PDF no contiene texto legible (posible imagen sin OCR).'}, status=400)

                    df_tablas = processor.extraer_tablas()
                    datospdf = processor.extraer_datos_cabecera()
                    processor.cerrar()
                    cant_muestras = int(len(df_tablas))
                    numero_acta = datospdf["numeroActa"]
                    cuitDeFuncionario = datospdf["cuitDeFuncionario"]
                    RENSPA = datospdf["RENSPA"]
                    motivo = datospdf["Motivo"]
                    submotivo = datospdf["SubMotivo"]
                    fechaToma = datospdf["FechaToma"]
                    especie = datospdf["Especie"]

                    resultadoLetra = int(resultadoLetra) if resultadoLetra and resultadoLetra.isdigit() else None
                    categoria = df_tablas["Categoria"].apply(lambda x: int(x) if pd.notna(x) and str(x).isdigit() else None)

                    df_temp = pd.DataFrame({
                        "Nro Informe": [nro_informe] * len(df_tablas),
                        "Nro Acta": [numero_acta] * len(df_tablas),
                        "RENSPA": [RENSPA] * len(df_tablas),
                        "Motivo": [motivo] * len(df_tablas),
                        "SubMotivo": [submotivo] * len(df_tablas),
                        "CUIT Funcionario": [cuitDeFuncionario] * len(df_tablas),
                        "Fecha de Toma": [fechaToma] * len(df_tablas),
                        "Fecha de Recepcion": [fechaRecepcion] * len(df_tablas),
                        "Especie": [especie] * len(df_tablas),
                        "Cantidad Muestras": [cant_muestras] * len(df_tablas),
                        "Rubro": [codigo_rubro] * len(df_tablas),
                        "Fecha Inicio": [fechaInicio] * len(df_tablas),
                        "Fecha Fin": [fechaFin] * len(df_tablas),
                        "Resultado Letra": [resultadoLetra] * len(df_tablas),
                        "Identificacion Muestra": df_tablas["Identificacion"].astype(str),
                        "Identificacion Interna Laboratorio": df_tablas["Nro_Tubo"],
                        "Tipo Identificacion": [12] * len(df_tablas),  # AIE usa tipo 12
                        "Observacion Muestra": [None] * len(df_tablas),
                        "Categoria": categoria,
                        "Sexo": df_tablas["Sexo"],
                        "Antigeno/Kit": [antigeno] * len(df_tablas),
                        "Marca Antigeno/Kit": [marcaAntigeno] * len(df_tablas),
                        "Lote": [lote] * len(df_tablas),
                        "Fecha Vencimiento Antigeno/Kit": [fechaVencimiento] * len(df_tablas),
                        "Estampilla": [estampilla] * len(df_tablas),
                        "Codigo DT": [codigoDT] * len(df_tablas),
                        "Observacion del Protocolo": [None] * len(df_tablas),
                        "Conclusion Protocolo": [conclusion] * len(df_tablas),
                    })

                    df_acta = pd.concat([df_acta, df_temp], ignore_index=True)

            # --- Conversión de tipos ---
            columnas_de_fecha = ["Fecha de Toma","Fecha de Recepcion","Fecha Inicio","Fecha Fin","Fecha Vencimiento Antigeno/Kit"]
            for columna in columnas_de_fecha:
                if columna in df_acta.columns:
                    df_acta[columna] = df_acta[columna].apply(formatear_fecha)

            columnas_enteros = ["Cantidad Muestras","Rubro"]
            for columna in columnas_enteros:
                if columna in df_acta.columns:
                    df_acta[columna] = pd.to_numeric(df_acta[columna], errors="coerce").astype("Int64")

            # --- Procesar DataFrame y generar JSON ---
            json_data = nexo_acta_json_aie(df_acta, request)

            fecha_actual = datetime.now().strftime("%d_%m_%y")
            nombre_archivo = f"AIE_Acta_Digital_{fecha_actual}.json"

            response = HttpResponse(
                json.dumps(json_data, ensure_ascii=False, indent=4, default=convertir_tipo),
                content_type='application/json'
            )
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            return response

        except Exception as e:
            import traceback
            error_msg = traceback.format_exc()
            print(f"❌ Error en la conversión:\n{error_msg}")
            return JsonResponse({'error': f'Ocurrió un error en la conversión: {str(e)}'}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)




#Crear el excel que usa el GRECERT para importacion de resultados
@csrf_exempt
def excel_GRECERT(request):
    if request.method == 'POST':
        try:
            columnas = [
                "numero", "identificación", "idTipoIdentificacion",
                "nroInternoLab","idCategoria","idEdad",
                "sexo","fechaVacunacion","antigenoKit",
                "marca","lote","vtoAntigeno","estampilla",
                "idResultadoLetra","resultadoNumero","idUnidadDeMedida","observación"
            ]

            acta = request.FILES.get("acta1")
            processor = PDFProcessor(acta)
            if not processor.es_valido():
                        logger.warning(" PDF sin texto legible (posible imagen sin OCR).")
                        return JsonResponse({'error': 'El PDF no contiene texto legible (posible imagen sin OCR).'}, status=400)

            df_tablas = processor.extraer_tablas()
            datospdf = processor.extraer_datos_cabecera()
            processor.cerrar()
            numero_acta = datospdf.get("numeroActa", "sin_numero")
            

            categoria = df_tablas["Categoria"].apply(
                lambda x: int(x) if pd.notna(x) and str(x).isdigit() else None
            ) if "Categoria" in df_tablas else None

            sexo = df_tablas["Sexo"] if "Sexo" in df_tablas else None

            df_acta = pd.DataFrame({
                "numero": df_tablas["Nro_Tubo"],
                "identificación": df_tablas["Identificacion"],
                "idTipoIdentificacion": None,
                "nroInternoLab": df_tablas["Nro_Tubo"], 
                "idCategoria": df_tablas["Categoria"],
                "idEdad": None,
                "sexo": sexo,
                "fechaVacunacion": None,
                "antigenoKit": request.POST.get("antigeno1"),
                "marca": request.POST.get("marcaAntigeno1"),
                "lote": request.POST.get("lote1"),
                "vtoAntigeno": request.POST.get("fechaVencimiento1"),
                "estampilla": request.POST.get("estampilla1"),
                "idResultadoLetra": None,
                "resultadoNumero": None,
                "idUnidadDeMedida": None,
                "observación": None
            })

            # Generar Excel en memoria
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_acta.to_excel(writer, sheet_name="Resultados", index=False, header=True)

            # Preparar respuesta HTTP
            nombre_archivo = f"Excel_GRECERT_ACTA{numero_acta}.xlsx"
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            return response

        except Exception as e:
            return JsonResponse({'error': f'Ocurrió un error en la conversión: {str(e)}'}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)

#Convierte las actas digitales en PDF de Brucelosis a EXCEL
@csrf_exempt
def convertir_Acta_bru(request):
    # --- Consulta de códigos de rubro ---
    codigo_rubro_bru = codigoRubro(request, 2)
    codigo_rubro_bru = codigo_rubro_bru[['codigo_rubro', 'tecnica']]
    tecnicas = codigo_rubro_bru['tecnica'].unique()

    codigo_rubro_tamiz = None
    codigo_rubro_confirmatoria = None

    if len(tecnicas) > 1:
        tamiz_df = codigo_rubro_bru.loc[
            codigo_rubro_bru['tecnica'].str.contains("BPAT", case=False, na=False),
            'codigo_rubro'
        ]
        confirmatoria_df = codigo_rubro_bru.loc[
            codigo_rubro_bru['tecnica'].str.contains("FPA|SAT Y 2-ME|FCT|I-ELISA", case=False, na=False),
            'codigo_rubro'
        ]
        if not tamiz_df.empty:
            codigo_rubro_tamiz = int(tamiz_df.iloc[0])
        if not confirmatoria_df.empty:
            codigo_rubro_confirmatoria = int(confirmatoria_df.iloc[0])
    elif len(tecnicas) == 1:
        unico_df = codigo_rubro_bru['codigo_rubro']
        codigo_rubro_tamiz = int(unico_df.iloc[0]) if not unico_df.empty else None

    # --- Procesamiento del POST ---
    if request.method == 'POST':
        try:
            columnas = [
                "Nro Informe","Nro Acta","RENSPA","Motivo","SubMotivo","CUIT Funcionario","Fecha de Toma","Fecha de Recepcion","Especie","Cantidad Muestras",
                "Rubro","Fecha Inicio","Fecha Fin",
                "Resultado Letra","Identificacion Muestra","Identificacion Interna Laboratorio", "Tipo Identificacion","Observacion Muestra",
                "Categoria","Sexo","Antigeno/Kit","Marca Antigeno/Kit","Lote",
                "Fecha Vencimiento Antigeno/Kit","Estampilla","Codigo DT",
                "Observacion del Protocolo","Conclusion Protocolo"
            ]
            df_acta = pd.DataFrame(columns=columnas)

            # --- Bucle para procesar hasta 4 actas ---
            for i in range(1, 5):
                acta = request.FILES.get(f"acta{i}")
                nro_informe = request.POST.get(f"informe{i}")
                fechaRecepcion = request.POST.get(f"fechaRecepcion{i}")
                fechaInicio = request.POST.get(f"fechaInicio{i}")
                fechaFin = request.POST.get(f"fechaFin{i}")
                fechaVencimiento = request.POST.get(f"fechaVencimiento{i}")
                resultadoLetra = request.POST.get(f"resultadoLetra{i}")
                antigeno = request.POST.get(f"antigeno{i}")
                marcaAntigeno = request.POST.get(f"marcaAntigeno{i}")
                lote = request.POST.get(f"lote{i}")
                estampilla = request.POST.get(f"estampilla{i}")
                codigoDT = request.POST.get(f"codigoDT{i}")
                conclusion = request.POST.get(f"conclusion{i}")

                if acta and nro_informe:
                    processor = PDFProcessor(acta)
                    if not processor.es_valido():
                        logger.warning(f"Acta {nro_informe}: PDF sin texto legible (posible imagen sin OCR).")
                        return JsonResponse({'error': 'El PDF no contiene texto legible (posible imagen sin OCR).'}, status=400)

                    df_tablas = processor.extraer_tablas()
                    datospdf = processor.extraer_datos_cabecera()
                    processor.cerrar()

                    cant_muestras = int(len(df_tablas))

                    numero_acta = datospdf["numeroActa"]
                    cuitDeFuncionario = datospdf["cuitDeFuncionario"]
                    RENSPA = datospdf["RENSPA"]
                    motivo = datospdf["Motivo"]
                    submotivo = datospdf["SubMotivo"]
                    fechaToma = datospdf["FechaToma"]
                    especie = datospdf["Especie"]

                    # --- Bloque principal (tamiz) ---
                    df_temp = pd.DataFrame({
                        "Nro Informe": [nro_informe] * len(df_tablas),
                        "Nro Acta": [numero_acta] * len(df_tablas),
                        "RENSPA": [RENSPA] * len(df_tablas),
                        "Motivo": [motivo] * len(df_tablas),
                        "SubMotivo": [submotivo] * len(df_tablas),
                        "CUIT Funcionario": [cuitDeFuncionario] * len(df_tablas),
                        "Fecha de Toma": [formatear_fecha(fechaToma)] * len(df_tablas),
                        "Fecha de Recepcion": [formatear_fecha(fechaRecepcion)] * len(df_tablas),
                        "Especie": [especie] * len(df_tablas),
                        "Cantidad Muestras": [cant_muestras] * len(df_tablas),
                        "Rubro": [codigo_rubro_tamiz] * len(df_tablas),
                        "Fecha Inicio": [formatear_fecha(fechaInicio)] * len(df_tablas),
                        "Fecha Fin": [formatear_fecha(fechaFin)] * len(df_tablas),
                        "Resultado Letra": [resultadoLetra] * len(df_tablas),
                        "Identificacion Muestra": df_tablas["Identificacion"].astype(str),
                        "Identificacion Interna Laboratorio": df_tablas["Nro_Tubo"],
                        "Tipo Identificacion": [1] * len(df_tablas),
                        "Observacion Muestra": None,
                        "Categoria": df_tablas["Categoria"],
                        "Sexo": df_tablas["Sexo"],
                        "Antigeno/Kit": [antigeno] * len(df_tablas),
                        "Marca Antigeno/Kit": [marcaAntigeno] * len(df_tablas),
                        "Lote": [lote] * len(df_tablas),
                        "Fecha Vencimiento Antigeno/Kit": [formatear_fecha(fechaVencimiento)] * len(df_tablas),
                        "Estampilla": [estampilla] * len(df_tablas),
                        "Codigo DT": [codigoDT] * len(df_tablas),
                        "Observacion del Protocolo": None,
                        "Conclusion Protocolo": [conclusion]*len(df_tablas),
                    })
                    df_acta = pd.concat([df_acta, df_temp], ignore_index=True)

                    # --- Bloque complementario (confirmatoria) ---
                    extra_info = request.POST.get(f"toggleComp{i}")
                    if extra_info == "on":
                        fechaInicioComp = request.POST.get(f"fechaInicioComp{i}")
                        fechaFinComp = request.POST.get(f"fechaFinComp{i}")
                        fechaVencimientoComp = request.POST.get(f"fechaVencimientoComp{i}")
                        resultadoLetraComp = request.POST.get(f"resultadoLetraComp{i}")
                        antigenoComp = request.POST.get(f"antigenoComp{i}")
                        marcaAntigenoComp = request.POST.get(f"marcaAntigenoComp{i}")
                        loteComp = request.POST.get(f"loteComp{i}")
                        estampillaComp = request.POST.get(f"estampillaComp{i}")

                        if any([fechaInicioComp, fechaFinComp, resultadoLetraComp, antigenoComp]):
                            df_comp = pd.DataFrame({
                                "Nro Informe": [nro_informe] * len(df_tablas),
                                "Nro Acta": [numero_acta] * len(df_tablas),
                                "RENSPA": [RENSPA] * len(df_tablas),
                                "Motivo": [motivo] * len(df_tablas),
                                "SubMotivo": [submotivo] * len(df_tablas),
                                "CUIT Funcionario": [cuitDeFuncionario] * len(df_tablas),
                                "Fecha de Toma": [formatear_fecha(fechaToma)] * len(df_tablas),
                                "Fecha de Recepcion": [formatear_fecha(fechaRecepcion)] * len(df_tablas),
                                "Especie": [especie] * len(df_tablas),
                                "Cantidad Muestras": [cant_muestras] * len(df_tablas),
                                "Rubro": [codigo_rubro_confirmatoria] * len(df_tablas),
                                "Fecha Inicio": [formatear_fecha(fechaInicioComp)] * len(df_tablas),
                                "Fecha Fin": [formatear_fecha(fechaFinComp)] * len(df_tablas),
                                "Resultado Letra": [resultadoLetraComp] * len(df_tablas),
                                "Identificacion Muestra": df_tablas["Identificacion"].astype(str),
                                "Identificacion Interna Laboratorio": df_tablas["Nro_Tubo"],
                                "Tipo Identificacion": [1] * len(df_tablas),
                                "Observacion Muestra": None,
                                "Categoria": df_tablas["Categoria"],
                                "Sexo": df_tablas["Sexo"],
                                "Antigeno/Kit": [antigenoComp] * len(df_tablas),
                                "Marca Antigeno/Kit": [marcaAntigenoComp] * len(df_tablas),
                                "Lote": [loteComp] * len(df_tablas),
                                "Fecha Vencimiento Antigeno/Kit": [formatear_fecha(fechaVencimientoComp)] * len(df_tablas),
                                "Estampilla": [estampillaComp] * len(df_tablas),
                                "Codigo DT": [codigoDT] * len(df_tablas),
                                "Observacion del Protocolo": None,
                                "Conclusion Protocolo": [conclusion] * len(df_tablas),
                            })
                            df_acta = pd.concat([df_acta, df_comp], ignore_index=True)

            # --- Conversión de tipos antes de exportar ---
            columnas_de_fecha = [
                "Fecha de Toma","Fecha de Recepcion","Fecha Inicio","Fecha Fin","Fecha Vencimiento Antigeno/Kit"
            ]
            for columna in columnas_de_fecha:
                if columna in df_acta.columns:
                    df_acta[columna] = df_acta[columna].apply(formatear_fecha)

            columnas_enteros = ["Cantidad Muestras","Rubro"]
            for columna in columnas_enteros:
                if columna in df_acta.columns:
                    df_acta[columna] = pd.to_numeric(df_acta[columna], errors="coerce").astype("Int64")

            # --- Generar Excel ---
            fecha_actual = datetime.now().strftime("%Y-%m-%d")
            nombre_archivo = f"Acta_Brucelosis_{fecha_actual}.xlsx"

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

            with pd.ExcelWriter(response, engine='openpyxl') as writer:
                df_acta.to_excel(writer, sheet_name="Acta Digital", index=False, header=True)

                # Pintar filas de confirmatoria en gris claro
                ws = writer.sheets["Acta Digital"]
                fill_gray = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

                col_rubro_idx = df_acta.columns.get_loc("Rubro") + 1  # +1 porque openpyxl es 1-based

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    rubro_cell = row[col_rubro_idx - 1]
                    if rubro_cell.value == codigo_rubro_confirmatoria:
                        for cell in row:
                            cell.fill = fill_gray

            return response

        except Exception as e:
            logger.error(f"Error en convertir_Acta_bru: {str(e)}")
            return JsonResponse({'error': f'Ocurrió un error en la conversión: {str(e)}'}, status=500)

    return JsonResponse({'error': 'No se pudo generar el archivo excel'}, status=405)

